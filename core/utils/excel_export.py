"""
Excel/CSV Export Utilities using Pandas
========================================

Professional Excel exports for accounting reports
"""

from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from datetime import datetime


def create_excel_response(filename='report.xlsx'):
    """Create an HTTP response for Excel file download"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def create_csv_response(filename='report.csv'):
    """Create an HTTP response for CSV file download"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_trial_balance_excel(report_data, form_data):
    """Export Trial Balance to Excel"""
    # Create Excel writer
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare data
    trial_balance_data = []
    for item in report_data['trial_balance']:
        trial_balance_data.append({
            'GL Code': item['account'].gl_code,
            'Account Name': item['account'].account_name,
            'Account Type': item['account'].account_type.get_name_display(),
            'Debit (₦)': float(item['debit']) if item['debit'] > 0 else 0,
            'Credit (₦)': float(item['credit']) if item['credit'] > 0 else 0,
        })

    # Create DataFrame
    df = pd.DataFrame(trial_balance_data)

    # Add totals row
    totals_row = pd.DataFrame([{
        'GL Code': '',
        'Account Name': 'TOTAL',
        'Account Type': '',
        'Debit (₦)': float(report_data['total_debits']),
        'Credit (₦)': float(report_data['total_credits']),
    }])
    df = pd.concat([df, totals_row], ignore_index=True)

    # Write to Excel
    df.to_excel(writer, sheet_name='Trial Balance', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Trial Balance']

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header styling
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_num, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Totals row styling
    totals_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    totals_font = Font(bold=True, size=11)
    last_row = len(df) + 1

    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=last_row, column=col_num)
        cell.fill = totals_fill
        cell.font = totals_font

    # Number formatting for currency columns
    for row in range(2, last_row + 1):
        worksheet.cell(row=row, column=4).number_format = '#,##0.00'  # Debit
        worksheet.cell(row=row, column=5).number_format = '#,##0.00'  # Credit

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 12  # GL Code
    worksheet.column_dimensions['B'].width = 40  # Account Name
    worksheet.column_dimensions['C'].width = 20  # Account Type
    worksheet.column_dimensions['D'].width = 18  # Debit
    worksheet.column_dimensions['E'].width = 18  # Credit

    # Add report header information
    worksheet.insert_rows(1, 3)
    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')
    worksheet.merge_cells('A3:E3')

    title_cell = worksheet['A1']
    title_cell.value = 'TRIAL BALANCE'
    title_cell.font = Font(bold=True, size=16, color='D97706')
    title_cell.alignment = Alignment(horizontal='center')

    period_cell = worksheet['A2']
    period_cell.value = f'Period: {report_data["date_from"].strftime("%B %d, %Y")} to {report_data["date_to"].strftime("%B %d, %Y")}'
    period_cell.alignment = Alignment(horizontal='center')

    balance_cell = worksheet['A3']
    balance_status = 'BALANCED ✓' if report_data['is_balanced'] else 'NOT BALANCED ✗'
    balance_cell.value = f'Status: {balance_status}'
    balance_cell.font = Font(bold=True, color='059669' if report_data['is_balanced'] else 'DC2626')
    balance_cell.alignment = Alignment(horizontal='center')

    # Save
    writer.close()
    output.seek(0)

    filename = f'trial_balance_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_profit_loss_excel(report_data, form_data):
    """Export Profit & Loss to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Income section
    income_data = []
    for item in report_data['income_items']:
        income_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Amount (₦)': float(item['amount']),
        })

    # Expense section
    expense_data = []
    for item in report_data['expense_items']:
        expense_data.append({
            'GL Code': item['account'].gl_code,
            'Account': item['account'].account_name,
            'Amount (₦)': float(item['amount']),
        })

    # Create DataFrames
    df_income = pd.DataFrame(income_data)
    df_expense = pd.DataFrame(expense_data)

    # Write to separate sheets
    df_income.to_excel(writer, sheet_name='Income', index=False)
    df_expense.to_excel(writer, sheet_name='Expenses', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Item': 'Total Income', 'Amount (₦)': float(report_data['total_income'])},
        {'Item': 'Total Expenses', 'Amount (₦)': float(report_data['total_expenses'])},
        {'Item': 'Net Profit/Loss', 'Amount (₦)': float(report_data['net_profit'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling (similar to trial balance)
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'profit_loss_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_general_ledger_excel(report_data, form_data):
    """Export General Ledger to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare transaction data
    transactions_data = []
    for txn in report_data['transactions']:
        transactions_data.append({
            'Date': txn['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Journal Number': txn['line'].journal_entry.journal_number,
            'Description': txn['line'].description,
            'Debit (₦)': float(txn['line'].debit_amount) if txn['line'].debit_amount > 0 else 0,
            'Credit (₦)': float(txn['line'].credit_amount) if txn['line'].credit_amount > 0 else 0,
            'Balance (₦)': float(txn['running_balance']),
        })

    df = pd.DataFrame(transactions_data)
    df.to_excel(writer, sheet_name='General Ledger', index=False)

    # Styling
    worksheet = writer.sheets['General Ledger']
    from openpyxl.styles import Font, PatternFill, Alignment

    # Add header info
    worksheet.insert_rows(1, 4)
    worksheet.merge_cells('A1:F1')
    worksheet.merge_cells('A2:F2')
    worksheet.merge_cells('A3:F3')

    worksheet['A1'] = 'GENERAL LEDGER'
    worksheet['A1'].font = Font(bold=True, size=16, color='D97706')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet['A2'] = f'{report_data["account"].gl_code} - {report_data["account"].account_name}'
    worksheet['A2'].font = Font(bold=True, size=12)
    worksheet['A2'].alignment = Alignment(horizontal='center')

    worksheet['A3'] = f'Period: {report_data["date_from"].strftime("%B %d, %Y")} to {report_data["date_to"].strftime("%B %d, %Y")}'
    worksheet['A3'].alignment = Alignment(horizontal='center')

    worksheet['A4'] = f'Opening Balance: ₦{report_data["opening_balance"]:,.2f}'
    worksheet['A4'].font = Font(bold=True)

    writer.close()
    output.seek(0)

    filename = f'general_ledger_{report_data["account"].gl_code}_{report_data["date_from"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_balance_sheet_excel(report_data, form_data):
    """Export Balance Sheet to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    def _acct_row(item):
        acct = item.get('account')
        return {
            'GL Code': acct.gl_code if acct else '',
            'Account': acct.account_name if acct else item.get('account_name', ''),
            'Balance (₦)': float(item['balance']),
        }

    # Assets section
    assets_data = [_acct_row(i) for i in report_data['assets']]

    # Liabilities section
    liabilities_data = [_acct_row(i) for i in report_data['liabilities']]

    # Equity section
    equity_data = [_acct_row(i) for i in report_data['equity']]

    # Create DataFrames
    df_assets = pd.DataFrame(assets_data)
    df_liabilities = pd.DataFrame(liabilities_data)
    df_equity = pd.DataFrame(equity_data)

    # Write to separate sheets
    df_assets.to_excel(writer, sheet_name='Assets', index=False)
    df_liabilities.to_excel(writer, sheet_name='Liabilities', index=False)
    df_equity.to_excel(writer, sheet_name='Equity', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Category': 'Total Assets', 'Amount (₦)': float(report_data['total_assets'])},
        {'Category': 'Total Liabilities', 'Amount (₦)': float(report_data['total_liabilities'])},
        {'Category': 'Total Equity', 'Amount (₦)': float(report_data['total_equity'])},
        {'Category': 'Total Liabilities + Equity', 'Amount (₦)': float(report_data['total_liabilities_equity'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'balance_sheet_{report_data["as_of_date"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_cash_flow_excel(report_data, form_data):
    """Export Cash Flow Statement to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Operating activities
    operating_data = []
    for item in report_data['operating_activities']:
        operating_data.append({
            'Date': item['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Description': item['line'].description,
            'Amount (₦)': float(item['amount']),
        })

    # Investing activities
    investing_data = []
    for item in report_data['investing_activities']:
        investing_data.append({
            'Date': item['line'].journal_entry.transaction_date.strftime('%Y-%m-%d'),
            'Description': item['line'].description,
            'Amount (₦)': float(item['amount']),
        })

    # Create DataFrames
    df_operating = pd.DataFrame(operating_data) if operating_data else pd.DataFrame(columns=['Date', 'Description', 'Amount (₦)'])
    df_investing = pd.DataFrame(investing_data) if investing_data else pd.DataFrame(columns=['Date', 'Description', 'Amount (₦)'])

    # Write to separate sheets
    df_operating.to_excel(writer, sheet_name='Operating Activities', index=False)
    df_investing.to_excel(writer, sheet_name='Investing Activities', index=False)

    # Create summary sheet
    summary_data = pd.DataFrame([
        {'Activity Type': 'Operating Activities', 'Total (₦)': float(report_data['operating_total'])},
        {'Activity Type': 'Investing Activities', 'Total (₦)': float(report_data['investing_total'])},
        {'Activity Type': 'Financing Activities', 'Total (₦)': float(report_data['financing_total'])},
        {'Activity Type': 'Net Cash Flow', 'Total (₦)': float(report_data['net_cash_flow'])},
    ])
    summary_data.to_excel(writer, sheet_name='Summary', index=False)

    # Apply styling
    from openpyxl.styles import Font, PatternFill, Alignment

    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]

        # Header styling
        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 18

    writer.close()
    output.seek(0)

    filename = f'cash_flow_{report_data["date_from"].strftime("%Y%m%d")}_{report_data["date_to"].strftime("%Y%m%d")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_transaction_audit_excel(report_data, form_data):
    """Export Transaction Audit Log to Excel"""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Prepare audit data
    audit_data_list = []
    for item in report_data['audit_data']:
        txn = item['transaction']
        journal_status = 'Yes' if item['has_journal'] else 'NO - MISSING ⚠️'

        audit_data_list.append({
            'Date': txn.transaction_date.strftime('%Y-%m-%d'),
            'Transaction Ref': txn.transaction_ref,
            'Type': txn.transaction_type,
            'Client': txn.client.get_full_name() if txn.client else 'N/A',
            'Amount (₦)': float(txn.amount),
            'Branch': txn.branch.name if txn.branch else 'N/A',
            'Has Journal Entry': journal_status,
        })

    df = pd.DataFrame(audit_data_list)
    df.to_excel(writer, sheet_name='Audit Log', index=False)

    # Styling
    worksheet = writer.sheets['Audit Log']
    from openpyxl.styles import Font, PatternFill, Alignment

    # Add header info
    worksheet.insert_rows(1, 3)
    worksheet.merge_cells('A1:G1')
    worksheet.merge_cells('A2:G2')

    worksheet['A1'] = 'TRANSACTION AUDIT LOG'
    worksheet['A1'].font = Font(bold=True, size=16, color='D97706')
    worksheet['A1'].alignment = Alignment(horizontal='center')

    worksheet['A2'] = f'Total Transactions: {report_data["total_transactions"]} | Missing Journal Entries: {report_data["missing_journal_count"]}'
    worksheet['A2'].font = Font(bold=True, color='DC2626' if report_data["missing_journal_count"] > 0 else '059669')
    worksheet['A2'].alignment = Alignment(horizontal='center')

    # Header row styling
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num in range(1, 8):
        cell = worksheet.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Highlight missing journal entries
    alert_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    for row in range(5, len(audit_data_list) + 5):
        status_cell = worksheet.cell(row=row, column=7)
        if 'MISSING' in str(status_cell.value):
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).fill = alert_fill

    # Column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 25

    writer.close()
    output.seek(0)

    filename = f'transaction_audit_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = create_excel_response(filename)
    response.write(output.read())
    return response


def export_to_csv(data, columns, filename='export.csv'):
    """Generic CSV export function"""
    df = pd.DataFrame(data, columns=columns)

    response = create_csv_response(filename)
    df.to_csv(response, index=False)
    return response


# ---------------------------------------------------------------------------
# Shared styling helper
# ---------------------------------------------------------------------------

def _style_sheet(worksheet, header_col_count, title=None, subtitle=None,
                 currency_cols=None, col_widths=None):
    """Apply consistent header styling and optional title rows to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment

    from openpyxl.utils import get_column_letter
    BRAND_ORANGE = 'D97706'
    HEADER_FILL = PatternFill(start_color=BRAND_ORANGE, end_color=BRAND_ORANGE, fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    TITLE_FONT  = Font(bold=True, size=14, color=BRAND_ORANGE)
    TOTAL_FILL  = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    insert_count = 0
    if title:
        insert_count += 1
    if subtitle:
        insert_count += 1

    if insert_count:
        worksheet.insert_rows(1, insert_count)
        row = 1
        col_letter = get_column_letter(header_col_count)
        if title:
            worksheet.merge_cells(f'A{row}:{col_letter}{row}')
            cell = worksheet[f'A{row}']
            cell.value = title
            cell.font = TITLE_FONT
            cell.alignment = Alignment(horizontal='center')
            row += 1
        if subtitle:
            worksheet.merge_cells(f'A{row}:{col_letter}{row}')
            cell = worksheet[f'A{row}']
            cell.value = subtitle
            cell.alignment = Alignment(horizontal='center')

    # Style header row (first row after title rows)
    header_row = insert_count + 1
    for col in range(1, header_col_count + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Style totals (last data row)
    last_row = worksheet.max_row
    if last_row > header_row + 1:
        for col in range(1, header_col_count + 1):
            cell = worksheet.cell(row=last_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)

    # Currency formatting
    if currency_cols:
        for row in range(header_row + 1, last_row + 1):
            for col in currency_cols:
                worksheet.cell(row=row, column=col).number_format = '#,##0.00'

    # Column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            worksheet.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# PAR Aging
# ---------------------------------------------------------------------------

def export_par_aging_excel(context):
    """Export PAR Aging Report to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    bucket_labels = {
        'current':  'Current (0 days)',
        'par_1_30': 'PAR 1-30 days',
        'par_31_60':'PAR 31-60 days',
        'par_61_90':'PAR 61-90 days',
        'par_90plus':'PAR 90+ days',
    }

    rows = []
    for key, label in bucket_labels.items():
        s = context['summary'][key]
        rows.append({
            'Bucket': label,
            'No. of Loans': s['count'],
            'Outstanding Balance (₦)': float(s['balance']),
            '% of Portfolio': s['pct'],
        })

    df_summary = pd.DataFrame(rows)
    totals = pd.DataFrame([{
        'Bucket': 'TOTAL',
        'No. of Loans': context['total_count'],
        'Outstanding Balance (₦)': float(context['total_balance']),
        '% of Portfolio': 100.0,
    }])
    df_summary = pd.concat([df_summary, totals], ignore_index=True)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)

    # Loan detail sheet
    detail_rows = []
    for key, label in bucket_labels.items():
        for loan, days in context['summary'][key]['items']:
            detail_rows.append({
                'Bucket': label,
                'Loan No.': loan.loan_number,
                'Client': loan.client.get_full_name(),
                'Branch': loan.branch.name if loan.branch else '',
                'Product': loan.loan_product.name if loan.loan_product else '',
                'Outstanding Balance (₦)': float(loan.outstanding_balance),
                'Days Overdue': days,
            })

    df_detail = pd.DataFrame(detail_rows) if detail_rows else pd.DataFrame(
        columns=['Bucket','Loan No.','Client','Branch','Product','Outstanding Balance (₦)','Days Overdue'])
    df_detail.to_excel(writer, sheet_name='Loan Detail', index=False)

    _style_sheet(writer.sheets['Summary'], 4,
                 title='PORTFOLIO AT RISK (PAR) AGING REPORT',
                 subtitle=f'As of {context["today"].strftime("%B %d, %Y")}',
                 currency_cols=[3], col_widths={'A':22,'B':16,'C':25,'D':18})
    _style_sheet(writer.sheets['Loan Detail'], 7,
                 currency_cols=[6], col_widths={'A':22,'B':18,'C':30,'D':20,'E':25,'F':25,'G':14})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'par_aging_{context["today"].strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Loan Officer Performance
# ---------------------------------------------------------------------------

def export_loan_officer_performance_excel(context):
    """Export Loan Officer Performance Report to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    rows = []
    for r in context['officers']:
        rows.append({
            'Officer': r['officer'].get_full_name(),
            'Branch': r['officer'].branch.name if r['officer'].branch else '',
            'Loans Disbursed': r['disbursed_count'],
            'Disbursed Amount (₦)': float(r['disbursed_amount']),
            'Repayments': r['repayment_count'],
            'Repayment Amount (₦)': float(r['repayment_amount']),
            'Active Clients': r['active_clients'],
            'Active Loans': r['active_loans'],
            'Overdue Loans': r['overdue_loans'],
        })

    totals = context['totals']
    rows.append({
        'Officer': 'TOTAL',
        'Branch': '',
        'Loans Disbursed': totals['disbursed_count'],
        'Disbursed Amount (₦)': float(totals['disbursed_amount']),
        'Repayments': totals['repayment_count'],
        'Repayment Amount (₦)': float(totals['repayment_amount']),
        'Active Clients': totals['active_clients'],
        'Active Loans': totals['active_loans'],
        'Overdue Loans': totals['overdue_loans'],
    })

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Performance', index=False)
    _style_sheet(writer.sheets['Performance'], 9,
                 title='LOAN OFFICER PERFORMANCE REPORT',
                 subtitle=f'Period: {context["date_from"]} to {context["date_to"]}',
                 currency_cols=[4, 6],
                 col_widths={'A':28,'B':20,'C':16,'D':22,'E':14,'F':22,'G':16,'H':14,'I':16})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'loan_officer_performance_{context["date_from"]}_{context["date_to"]}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Savings Maturity
# ---------------------------------------------------------------------------

def export_savings_maturity_excel(context):
    """Export Savings Maturity Report to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    def _rows(accounts):
        return [{
            'Account No.': r['account'].account_number,
            'Client': r['account'].client.get_full_name(),
            'Branch': r['account'].branch.name if r['account'].branch else '',
            'Product': r['account'].savings_product.name,
            'Balance (₦)': float(r['account'].balance),
            'Maturity Date': r['account'].maturity_date.strftime('%Y-%m-%d') if r['account'].maturity_date else '',
            'Days Remaining': r['days_left'],
        } for r in accounts]

    cols = ['Account No.','Client','Branch','Product','Balance (₦)','Maturity Date','Days Remaining']

    for label, data in [('Overdue', context['overdue']),
                        ('This Month', context['this_month']),
                        ('Next Month', context['next_month']),
                        ('Later', context['later'])]:
        df = pd.DataFrame(_rows(data)) if data else pd.DataFrame(columns=cols)
        df.to_excel(writer, sheet_name=label, index=False)
        _style_sheet(writer.sheets[label], 7,
                     currency_cols=[5],
                     col_widths={'A':18,'B':28,'C':18,'D':28,'E':18,'F':14,'G':16})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'savings_maturity_{context["today"].strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Client Transactions (detail page tab)
# ---------------------------------------------------------------------------

def export_client_transactions_excel(client, transactions):
    """Export all transactions for a client (client detail → Transactions tab)."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    rows = []
    for t in transactions:
        rows.append({
            'Date': t.transaction_date.strftime('%Y-%m-%d %H:%M') if hasattr(t.transaction_date, 'strftime') else str(t.transaction_date),
            'Reference': t.transaction_ref or '',
            'Type': t.get_transaction_type_display(),
            'Description': t.description or '',
            'Amount (₦)': float(t.amount),
            'Status': t.get_status_display() if hasattr(t, 'get_status_display') else t.status,
            'Processed By': t.processed_by.get_full_name() if t.processed_by else '',
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date','Reference','Type','Description','Amount (₦)','Status','Processed By'])
    df.to_excel(writer, sheet_name='Transactions', index=False)
    _style_sheet(writer.sheets['Transactions'], 7,
                 title=f'TRANSACTIONS — {client.get_full_name()} ({client.client_id})',
                 currency_cols=[5],
                 col_widths={'A':20,'B':20,'C':22,'D':40,'E':18,'F':14,'G':24})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'transactions_{client.client_id}_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Client Statement
# ---------------------------------------------------------------------------

def export_client_statement_excel(client, transactions, date_from, date_to,
                                   total_in, total_out, net_position):
    """Export client account statement to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    _OUTFLOW = {'loan_disbursement', 'withdrawal'}

    rows = []
    for t in transactions:
        rows.append({
            'Date': t.transaction_date.strftime('%Y-%m-%d') if hasattr(t.transaction_date, 'strftime') else str(t.transaction_date),
            'Reference': t.transaction_ref or '',
            'Type': t.get_transaction_type_display(),
            'Description': t.description or '',
            'Money In (₦)': float(t.amount) if t.transaction_type not in _OUTFLOW else 0.0,
            'Money Out (₦)': float(t.amount) if t.transaction_type in _OUTFLOW else 0.0,
        })

    # Summary rows
    rows.append({'Date':'','Reference':'','Type':'','Description':'TOTAL IN','Money In (₦)':float(total_in),'Money Out (₦)':''})
    rows.append({'Date':'','Reference':'','Type':'','Description':'TOTAL OUT','Money In (₦)':'','Money Out (₦)':float(total_out)})
    rows.append({'Date':'','Reference':'','Type':'','Description':'NET POSITION','Money In (₦)':float(net_position),'Money Out (₦)':''})

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Statement', index=False)

    date_range = f'{date_from.strftime("%B %d, %Y")} to {date_to.strftime("%B %d, %Y")}'
    _style_sheet(writer.sheets['Statement'], 6,
                 title=f'ACCOUNT STATEMENT — {client.get_full_name()} ({client.client_id})',
                 subtitle=f'Period: {date_range}',
                 currency_cols=[5, 6],
                 col_widths={'A':14,'B':20,'C':22,'D':40,'E':18,'F':18})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'statement_{client.client_id}_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Subsidiary Ledger
# ---------------------------------------------------------------------------

def export_subsidiary_ledger_excel(client, lines_qs, date_from, date_to,
                                    account_summary, totals):
    """Export subsidiary ledger to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # Detail sheet
    rows = []
    for line in lines_qs:
        je = line.journal_entry
        rows.append({
            'Date': je.transaction_date.strftime('%Y-%m-%d'),
            'Journal No.': je.journal_number,
            'GL Code': line.account.gl_code,
            'Account': line.account.account_name,
            'Description': line.description or je.description or '',
            'Debit (₦)': float(line.debit_amount) if line.debit_amount else 0.0,
            'Credit (₦)': float(line.credit_amount) if line.credit_amount else 0.0,
            'Branch': je.branch.name if je.branch else '',
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date','Journal No.','GL Code','Account','Description','Debit (₦)','Credit (₦)','Branch'])
    df.to_excel(writer, sheet_name='Ledger Lines', index=False)
    _style_sheet(writer.sheets['Ledger Lines'], 8,
                 title=f'SUBSIDIARY LEDGER — {client.get_full_name()} ({client.client_id})',
                 subtitle=f'Period: {date_from.strftime("%B %d, %Y")} to {date_to.strftime("%B %d, %Y")}',
                 currency_cols=[6, 7],
                 col_widths={'A':14,'B':18,'C':10,'D':30,'E':40,'F':16,'G':16,'H':18})

    # Account summary sheet
    summary_rows = [{
        'GL Code': s['account__gl_code'],
        'Account': s['account__account_name'],
        'Total Debit (₦)': float(s['total_debit'] or 0),
        'Total Credit (₦)': float(s['total_credit'] or 0),
    } for s in account_summary]
    if totals:
        summary_rows.append({
            'GL Code': '', 'Account': 'GRAND TOTAL',
            'Total Debit (₦)': float(totals.get('total_debit') or 0),
            'Total Credit (₦)': float(totals.get('total_credit') or 0),
        })

    df_sum = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(
        columns=['GL Code','Account','Total Debit (₦)','Total Credit (₦)'])
    df_sum.to_excel(writer, sheet_name='Account Summary', index=False)
    _style_sheet(writer.sheets['Account Summary'], 4,
                 currency_cols=[3, 4],
                 col_widths={'A':12,'B':35,'C':20,'D':20})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'subsidiary_ledger_{client.client_id}_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Client List
# ---------------------------------------------------------------------------

def export_client_list_excel(clients_qs):
    """Export full client list with all key details to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    rows = []
    for c in clients_qs.select_related('branch', 'assigned_staff', 'group'):
        rows.append({
            'Client ID': c.client_id,
            'First Name': c.first_name,
            'Last Name': c.last_name,
            'Gender': c.get_gender_display() if hasattr(c, 'get_gender_display') else (c.gender or ''),
            'Date of Birth': c.date_of_birth.strftime('%Y-%m-%d') if c.date_of_birth else '',
            'Phone': c.phone or '',
            'Alternate Phone': c.alternate_phone or '',
            'Email': c.email or '',
            'Address': c.address or '',
            'City': c.city or '',
            'State': c.state or '',
            'Branch': c.branch.name if c.branch else '',
            'Assigned Staff': c.assigned_staff.get_full_name() if c.assigned_staff else '',
            'Group': c.group.name if c.group else '',
            'Group Role': c.get_group_role_display() if c.group_role else '',
            'Occupation': c.occupation or '',
            'Business Name': c.business_name or '',
            'Monthly Income (₦)': float(c.monthly_income) if c.monthly_income else '',
            'ID Type': c.get_id_type_display() if c.id_type else '',
            'ID Number': c.id_number or '',
            'BVN': c.bvn or '',
            'Bank Name': c.bank_name or '',
            'Account Number': c.account_number or '',
            'Status': 'Active' if c.is_active else 'Inactive',
            'Approval Status': c.get_approval_status_display() if hasattr(c, 'get_approval_status_display') else (c.approval_status or ''),
            'Date Joined': c.created_at.strftime('%Y-%m-%d') if c.created_at else '',
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    df.to_excel(writer, sheet_name='Clients', index=False)
    _style_sheet(writer.sheets['Clients'], 26,
                 title='CLIENT LIST — SEASHORE MICROFINANCE',
                 subtitle=f'Exported on {datetime.now().strftime("%B %d, %Y at %H:%M")}',
                 currency_cols=[18],
                 col_widths={
                    'A':14,'B':18,'C':18,'D':10,'E':14,'F':16,'G':16,'H':28,
                    'I':35,'J':16,'K':16,'L':20,'M':24,'N':20,'O':14,'P':20,
                    'Q':25,'R':18,'S':14,'T':20,'U':16,'V':18,'W':18,'X':12,
                    'Y':18,'Z':14,
                 })

    writer.close()
    output.seek(0)
    response = create_excel_response(f'clients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    response.write(output.read())
    return response


# ---------------------------------------------------------------------------
# Savings Transactions
# ---------------------------------------------------------------------------

def export_savings_transactions_excel(postings):
    """Export savings deposit/withdrawal postings to Excel."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    rows = []
    for p in postings:
        is_deposit = hasattr(p, 'deposit_ref') or p.__class__.__name__ == 'SavingsDepositPosting'
        rows.append({
            'Date': p.submitted_at.strftime('%Y-%m-%d %H:%M') if p.submitted_at else '',
            'Reference': getattr(p, 'deposit_ref', None) or getattr(p, 'withdrawal_ref', '') or '',
            'Type': 'Deposit' if is_deposit else 'Withdrawal',
            'Client': p.client.get_full_name() if p.client else '',
            'Account No.': p.savings_account.account_number if p.savings_account else '',
            'Amount (₦)': float(p.amount),
            'Branch': p.branch.name if p.branch else '',
            'Status': p.get_status_display() if hasattr(p, 'get_status_display') else p.status,
            'Submitted By': p.submitted_by.get_full_name() if p.submitted_by else '',
            'Reviewed By': p.reviewed_by.get_full_name() if p.reviewed_by else '',
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date','Reference','Type','Client','Account No.','Amount (₦)',
                 'Branch','Status','Submitted By','Reviewed By'])
    df.to_excel(writer, sheet_name='Transactions', index=False)
    _style_sheet(writer.sheets['Transactions'], 10,
                 title='SAVINGS TRANSACTIONS',
                 subtitle=f'Exported on {datetime.now().strftime("%B %d, %Y at %H:%M")}',
                 currency_cols=[6],
                 col_widths={'A':20,'B':20,'C':12,'D':28,'E':18,'F':18,
                             'G':18,'H':14,'I':24,'J':24})

    writer.close()
    output.seek(0)
    response = create_excel_response(f'savings_transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    response.write(output.read())
    return response
